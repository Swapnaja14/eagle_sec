import os

from django.http import FileResponse, Http404
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from . import services
# PDF generation temporarily disabled - reportlab not installed
# from utils.pdf import generate_analytics_report_pdf

class EmployeeProgressView(APIView):
    """
    GET /analytics/employee/{id}
    Returns training completion stats and average assessment score.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, employee_id):
        tenant = None if request.user.role == "superadmin" else request.user.tenant
        data = services.get_employee_progress(employee_id, tenant=tenant)
        if not data:
            return Response(
                {"detail": f"Employee {employee_id} not found or not a trainee."},
                status=404,
            )
        return Response(data)


class TrainerPerformanceView(APIView):
    """
    GET /analytics/trainer/{id}
    Returns average trainee scores and feedback ratings for a trainer.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, trainer_id):
        tenant = None if request.user.role == "superadmin" else request.user.tenant
        data = services.get_trainer_performance(trainer_id, tenant=tenant)
        if not data:
            return Response(
                {"detail": f"Trainer {trainer_id} not found or not an instructor."},
                status=404,
            )
        return Response(data)


class OverallSummaryView(APIView):
    """
    GET /analytics/summary
    Returns platform-wide aggregated stats.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        tenant = None if request.user.role == "superadmin" else request.user.tenant
        return Response(services.get_overall_summary(tenant=tenant))


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def analytics_report(request):
    """
    GET /analytics/report
    Generates and streams a downloadable PDF analytics report.
    NOTE: PDF generation temporarily disabled - reportlab not installed
    """
    return Response({
        "detail": "PDF generation temporarily unavailable. Install reportlab and Pillow to enable this feature.",
        "install_command": "pip install reportlab Pillow"
    }, status=503)


class GapAnalysisView(APIView):
    """
    GET /analytics/gap-analysis/
    Returns skill gap data: required vs current proficiency per course,
    department gap scores, radar data, and summary KPIs.
    Query params: department (optional)
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        tenant = None if request.user.role == "superadmin" else request.user.tenant
        department = request.query_params.get("department", "")
        data = services.get_gap_analysis(tenant=tenant, department=department)
        return Response(data)


import csv
import io
import datetime
from django.http import HttpResponse

class BulkExportView(APIView):
    """
    POST /api/analytics/bulk-export/
    Generates a report file (CSV, Excel, or PDF) based on the given report type
    and filter options, returned as a downloadable file.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        report_type = request.data.get('reportType', 'training_history')
        fmt = request.data.get('format', 'csv')
        client_id = request.data.get('clientId', '')
        site_id = request.data.get('siteId', '')
        department = request.data.get('department', '')
        date_from = request.data.get('dateFrom', '')
        date_to = request.data.get('dateTo', '')

        tenant = None if request.user.role == 'superadmin' else request.user.tenant

        # Build the data rows based on report type
        rows, headers = self._get_report_data(report_type, tenant, client_id, site_id, department, date_from, date_to)

        filename = f"{report_type}_{datetime.date.today().isoformat()}"

        if fmt == 'csv':
            return self._generate_csv(headers, rows, filename)
        elif fmt == 'excel':
            return self._generate_excel(headers, rows, filename)
        elif fmt == 'pdf':
            return self._generate_pdf(headers, rows, filename, report_type)
        else:
            return Response({"error": "Unsupported format"}, status=400)

    def _get_report_data(self, report_type, tenant, client_id, site_id, department, date_from, date_to):
        from django.contrib.auth import get_user_model
        from assessments.models import Submission
        from attendance.models import Attendance
        from dashboard.models import TrainingSession

        User = get_user_model()

        if report_type == 'training_history':
            headers = ['Session ID', 'Topic', 'Type', 'Status', 'Date', 'Trainer']
            qs = TrainingSession.objects.filter(is_active=True).select_related('trainer')
            if tenant:
                qs = qs.filter(tenant=tenant)
            if date_from:
                qs = qs.filter(date_time__date__gte=date_from)
            if date_to:
                qs = qs.filter(date_time__date__lte=date_to)
            rows = []
            for s in qs.order_by('-date_time'):
                trainer_name = ''
                if s.trainer:
                    trainer_name = f"{s.trainer.first_name} {s.trainer.last_name}".strip() or s.trainer.username
                rows.append([s.id, s.topic, s.session_type, s.status, str(s.date_time), trainer_name])

        elif report_type == 'quiz_results':
            headers = ['Submission ID', 'Quiz', 'User', 'Score', 'Percentage', 'Passed', 'Status', 'Date']
            qs = Submission.objects.select_related('quiz', 'user').order_by('-created_at')
            if department:
                qs = qs.filter(user__department__iexact=department)
            if date_from:
                qs = qs.filter(created_at__date__gte=date_from)
            if date_to:
                qs = qs.filter(created_at__date__lte=date_to)
            rows = []
            for sub in qs:
                user_name = f"{sub.user.first_name} {sub.user.last_name}".strip() or sub.user.username
                rows.append([sub.id, sub.quiz.title, user_name, sub.score, f"{sub.percentage}%", 'Yes' if sub.passed else 'No', sub.status, str(sub.submitted_at or sub.created_at)])

        elif report_type == 'psara_compliance':
            from certificates.models import IssuedCertificate
            headers = ['Employee', 'Department', 'Course', 'Issued At', 'Certificate ID']
            qs = IssuedCertificate.objects.select_related('employee', 'course').order_by('-issued_at')
            rows = []
            for c in qs:
                emp_name = f"{c.employee.first_name} {c.employee.last_name}".strip() or c.employee.username
                rows.append([emp_name, c.employee.department, c.course.display_name, str(c.issued_at), c.id])

        elif report_type == 'analytics':
            headers = ['Employee', 'Department', 'Assigned', 'Completed', 'Completion %', 'Avg Score']
            data = services.get_all_employee_progress(tenant=tenant)
            if department:
                data = [d for d in data if d.get('department', '').lower() == department.lower()]
            rows = []
            for d in data:
                rows.append([d['employee_name'], d['department'], d['total_assigned'], d['total_completed'], f"{d['completion_percentage']}%", d.get('average_assessment_score', 'N/A')])

        elif report_type == 'gap_analysis':
            gap_data = services.get_gap_analysis(tenant=tenant, department=department)
            headers = ['Skill/Course', 'Required %', 'Current %', 'Gap %']
            rows = []
            for sg in gap_data.get('skill_gaps', []):
                rows.append([sg['skill'], sg['required'], sg['current'], sg['gap']])
        else:
            headers = ['Info']
            rows = [['Unknown report type']]

        return rows, headers

    def _generate_csv(self, headers, rows, filename):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        return response

    def _generate_excel(self, headers, rows, filename):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Report'

        # Header styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response

    def _generate_pdf(self, headers, rows, filename, report_type):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = []

        # Title
        title = report_type.replace('_', ' ').title() + ' Report'
        elements.append(Paragraph(title, styles['Title']))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Generated on {datetime.date.today().isoformat()}", styles['Normal']))
        elements.append(Spacer(1, 24))

        # Table
        table_data = [headers] + [[str(v) for v in row] for row in rows]

        if table_data:
            t = Table(table_data, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ]))
            elements.append(t)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return response
