"""
Bulk User Upload Service
Handles CSV/Excel file parsing and bulk user creation with validation.
"""
import csv
import io
import re
from typing import List, Dict, Any, Tuple
from django.contrib.auth import get_user_model
from django.db import transaction
from .models import Tenant, Client, Site

User = get_user_model()

# Required CSV headers
REQUIRED_HEADERS = [
    'employee_id', 'first_name', 'last_name', 'email', 
    'department', 'designation', 'client', 'site', 'role'
]

# Role mapping: CSV value -> database value
ROLE_MAPPING = {
    'trainee': 'trainee',
    'guard': 'trainee',
    'trainer': 'instructor',
    'instructor': 'instructor',
    'admin': 'admin',
    'manager': 'admin',
    'superadmin': 'superadmin',
    'super admin': 'superadmin',
}


def normalize_header(header: str) -> str:
    """Normalize CSV header to lowercase with underscores."""
    return header.strip().lower().replace(' ', '_')


def validate_email(email: str) -> bool:
    """Validate email format."""
    if not email:
        return False
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, email))


def validate_role(role: str) -> Tuple[bool, str]:
    """Validate and normalize role."""
    role_lower = role.strip().lower()
    if role_lower in ROLE_MAPPING:
        return True, ROLE_MAPPING[role_lower]
    return False, ''


def parse_csv_file(file_content: bytes, tenant: Tenant = None) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Parse CSV file and validate rows.
    Returns: (validated_rows, errors)
    """
    try:
        # Decode file content
        content = file_content.decode('utf-8-sig')  # Handle BOM
        csv_file = io.StringIO(content)
        reader = csv.DictReader(csv_file)
        
        # Normalize headers
        if reader.fieldnames:
            reader.fieldnames = [normalize_header(h) for h in reader.fieldnames]
        
        # Check required headers
        missing_headers = set(REQUIRED_HEADERS) - set(reader.fieldnames or [])
        if missing_headers:
            return [], [f"Missing required columns: {', '.join(missing_headers)}"]
        
        validated_rows = []
        errors = []
        seen_employee_ids = set()
        seen_emails = set()
        
        for idx, row in enumerate(reader, start=2):  # Start at 2 (header is row 1)
            row_errors = []
            
            # Extract and validate fields
            employee_id = row.get('employee_id', '').strip()
            first_name = row.get('first_name', '').strip()
            last_name = row.get('last_name', '').strip()
            email = row.get('email', '').strip()
            department = row.get('department', '').strip()
            designation = row.get('designation', '').strip()
            client_name = row.get('client', '').strip()
            site_name = row.get('site', '').strip()
            role = row.get('role', '').strip()
            
            # Validate required fields
            if not employee_id:
                row_errors.append("Employee ID is required")
            elif employee_id in seen_employee_ids:
                row_errors.append(f"Duplicate Employee ID: {employee_id}")
            else:
                seen_employee_ids.add(employee_id)
                # Check if employee_id already exists in database
                if User.objects.filter(username=employee_id).exists():
                    row_errors.append(f"Employee ID {employee_id} already exists in system")
            
            if not first_name:
                row_errors.append("First name is required")
            
            if not last_name:
                row_errors.append("Last name is required")
            
            if not email:
                row_errors.append("Email is required")
            elif not validate_email(email):
                row_errors.append("Invalid email format")
            elif email in seen_emails:
                row_errors.append(f"Duplicate email: {email}")
            else:
                seen_emails.add(email)
                # Check if email already exists in database
                if User.objects.filter(email=email).exists():
                    row_errors.append(f"Email {email} already exists in system")
            
            if not role:
                row_errors.append("Role is required")
            else:
                valid_role, normalized_role = validate_role(role)
                if not valid_role:
                    row_errors.append(f"Invalid role: {role}. Must be one of: trainee, guard, trainer, instructor, admin, manager, superadmin")
                else:
                    role = normalized_role
            
            # Build validated row
            validated_row = {
                'employee_id': employee_id,
                'first_name': first_name,
                'last_name': last_name,
                'email': email,
                'department': department,
                'designation': designation,
                'client_name': client_name,
                'site_name': site_name,
                'role': role,
                'row_number': idx,
                'status': 'error' if row_errors else 'valid',
                'errors': row_errors,
            }
            
            validated_rows.append(validated_row)
        
        return validated_rows, errors
        
    except UnicodeDecodeError:
        return [], ["File encoding error. Please save the CSV file as UTF-8."]
    except Exception as e:
        return [], [f"Error parsing CSV file: {str(e)}"]


def parse_excel_file(file_content: bytes, tenant: Tenant = None) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Parse Excel file and validate rows.
    Returns: (validated_rows, errors)
    """
    try:
        import openpyxl
        from openpyxl import load_workbook
    except ImportError:
        return [], ["openpyxl library not installed. Please install it to support Excel files."]
    
    try:
        # Load workbook from bytes
        workbook = load_workbook(io.BytesIO(file_content), read_only=True)
        sheet = workbook.active
        
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(normalize_header(str(cell.value)))
        
        # Check required headers
        missing_headers = set(REQUIRED_HEADERS) - set(headers)
        if missing_headers:
            return [], [f"Missing required columns: {', '.join(missing_headers)}"]
        
        validated_rows = []
        seen_employee_ids = set()
        seen_emails = set()
        
        # Process data rows (skip header)
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue
            
            # Create dict from row
            row_dict = {}
            for i, header in enumerate(headers):
                if i < len(row):
                    row_dict[header] = str(row[i]).strip() if row[i] is not None else ''
            
            row_errors = []
            
            # Extract and validate fields
            employee_id = row_dict.get('employee_id', '').strip()
            first_name = row_dict.get('first_name', '').strip()
            last_name = row_dict.get('last_name', '').strip()
            email = row_dict.get('email', '').strip()
            department = row_dict.get('department', '').strip()
            designation = row_dict.get('designation', '').strip()
            client_name = row_dict.get('client', '').strip()
            site_name = row_dict.get('site', '').strip()
            role = row_dict.get('role', '').strip()
            
            # Validate required fields
            if not employee_id:
                row_errors.append("Employee ID is required")
            elif employee_id in seen_employee_ids:
                row_errors.append(f"Duplicate Employee ID: {employee_id}")
            else:
                seen_employee_ids.add(employee_id)
                if User.objects.filter(username=employee_id).exists():
                    row_errors.append(f"Employee ID {employee_id} already exists in system")
            
            if not first_name:
                row_errors.append("First name is required")
            
            if not last_name:
                row_errors.append("Last name is required")
            
            if not email:
                row_errors.append("Email is required")
            elif not validate_email(email):
                row_errors.append("Invalid email format")
            elif email in seen_emails:
                row_errors.append(f"Duplicate email: {email}")
            else:
                seen_emails.add(email)
                if User.objects.filter(email=email).exists():
                    row_errors.append(f"Email {email} already exists in system")
            
            if not role:
                row_errors.append("Role is required")
            else:
                valid_role, normalized_role = validate_role(role)
                if not valid_role:
                    row_errors.append(f"Invalid role: {role}")
                else:
                    role = normalized_role
            
            # Build validated row
            validated_row = {
                'employee_id': employee_id,
                'first_name': first_name,
                'last_name': last_name,
                'email': email,
                'department': department,
                'designation': designation,
                'client_name': client_name,
                'site_name': site_name,
                'role': role,
                'row_number': idx,
                'status': 'error' if row_errors else 'valid',
                'errors': row_errors,
            }
            
            validated_rows.append(validated_row)
        
        return validated_rows, []
        
    except Exception as e:
        return [], [f"Error parsing Excel file: {str(e)}"]


@transaction.atomic
def bulk_create_users(validated_rows: List[Dict[str, Any]], tenant: Tenant = None) -> Dict[str, Any]:
    """
    Create users in bulk from validated rows.
    Returns: {created_count, skipped_count, errors}
    """
    created_count = 0
    skipped_count = 0
    errors = []
    
    # Filter only valid rows
    valid_rows = [row for row in validated_rows if row['status'] == 'valid']
    
    for row in valid_rows:
        try:
            # Create user with employee_id as username
            user = User.objects.create_user(
                username=row['employee_id'],
                email=row['email'],
                first_name=row['first_name'],
                last_name=row['last_name'],
                role=row['role'],
                department=row['department'] if row['department'] else '',
                tenant=tenant,
                password=f"{row['employee_id']}@123",  # Default password
            )
            created_count += 1
            
        except Exception as e:
            skipped_count += 1
            errors.append(f"Row {row['row_number']}: {str(e)}")
    
    return {
        'created_count': created_count,
        'skipped_count': skipped_count,
        'errors': errors,
    }
